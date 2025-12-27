"""
GASB 75 Disclosure Population Module
=====================================
Populates the GASB 75 footnote tables Excel workbook with valuation results.

Data Flow Architecture:
1. Valuation engine produces results → ProVal1 tab
2. ProVal1 feeds → Net OPEB, 1%-+, RSI, OPEB Exp & Def
3. Prior year file provides → RSI historical columns, Table7AmortDeferred yellow cells, 
   OPEB Exp & Def prior year amortization values

ProVal1 Column Structure:
- B: BOY liability at prior year EOY discount rate
- C: BOY liability at current BOY discount rate (for assumption change calc)
- D: EOY liability - baseline
- E: EOY liability - discount rate +1%
- F: EOY liability - discount rate -1%
- G: EOY liability - trend baseline (same as D)
- H: EOY liability - trend +1%
- I: EOY liability - trend -1%

Key ProVal1 Rows:
- Row 6: Total Number of Actives
- Row 8: Number Inactive (retirees)
- Row 17: Total Salary (covered payroll)
- Row 19: EAN Acctg Liab (Total OPEB Liability)
- Row 38: EAN Acctg Normal Cost (Service Cost)
- Row 74: Avg Exp Service Lives (ARSL)
- Row 88: Accounting Interest Rate (discount rate)
"""

from dataclasses import dataclass, field
from typing import Dict, List, Optional, Any
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.formatting import ConditionalFormattingList
import copy


@dataclass
class ValuationResults:
    """Container for all valuation engine outputs needed for GASB 75 disclosure."""
    
    # Identification
    client_name: str
    measurement_date: date  # EOY measurement date (e.g., 9/30/2025)
    prior_measurement_date: date  # BOY (e.g., 9/30/2024)
    
    # Census counts
    active_count: int
    retiree_count: int
    covered_payroll: float
    
    # Liabilities - BOY
    tol_boy_old_rate: float  # BOY liability at prior year EOY rate
    tol_boy_new_rate: float  # BOY liability restated at current BOY rate
    
    # Liabilities - EOY (baseline and sensitivities)
    tol_eoy_baseline: float
    tol_eoy_disc_plus_1: float  # Discount rate +1%
    tol_eoy_disc_minus_1: float  # Discount rate -1%
    tol_eoy_trend_plus_1: float  # Healthcare trend +1%
    tol_eoy_trend_minus_1: float  # Healthcare trend -1%
    
    # Costs
    service_cost: float  # Normal cost at BOY rate
    
    # Assumptions
    discount_rate_boy: float  # Prior measurement date rate
    discount_rate_eoy: float  # Current measurement date rate
    avg_remaining_service_life: float
    
    # Optional detailed census (for ProVal1 completeness)
    active_count_fully_eligible: int = 0
    total_salary_fully_eligible: float = 0.0
    
    # Benefit payments (if any)
    benefit_payments: float = 0.0


@dataclass  
class PriorYearData:
    """Data extracted from prior year's GASB 75 disclosure file."""
    
    # RSI historical values (columns B through H for years 2018-2024)
    rsi_service_cost: List[float] = field(default_factory=list)
    rsi_interest: List[float] = field(default_factory=list)
    rsi_benefit_changes: List[Any] = field(default_factory=list)
    rsi_experience: List[float] = field(default_factory=list)
    rsi_assumption_changes: List[float] = field(default_factory=list)
    rsi_benefit_payments: List[float] = field(default_factory=list)
    rsi_net_change: List[float] = field(default_factory=list)
    rsi_tol_beginning: List[float] = field(default_factory=list)
    rsi_tol_ending: List[float] = field(default_factory=list)
    rsi_covered_payroll: List[float] = field(default_factory=list)
    rsi_liability_pct: List[float] = field(default_factory=list)
    rsi_discount_rates: List[Any] = field(default_factory=list)
    rsi_mortality: List[str] = field(default_factory=list)
    rsi_trend: List[str] = field(default_factory=list)
    
    # OPEB Exp & Def - prior year amortization values
    opeb_assumption_change_prior: float = 0.0  # B7 in prior year -> B7 current
    opeb_experience_prior: float = 0.0  # B24 in prior year -> B24 current
    
    # Table7AmortDeferred - yellow cells (prior year totals and service lives)
    amort_experience_years: Dict[int, Dict[str, Any]] = field(default_factory=dict)
    amort_assumption_years: Dict[int, Dict[str, Any]] = field(default_factory=dict)


def extract_prior_year_data(prior_year_path: str) -> PriorYearData:
    """
    Extract required data from prior year's GASB 75 disclosure file.
    
    Args:
        prior_year_path: Path to prior year's GASB 75 Excel file
        
    Returns:
        PriorYearData with all values needed for current year disclosure
    """
    wb = load_workbook(prior_year_path, data_only=True)
    data = PriorYearData()
    
    # RSI Tab - extract columns B through H (years 2018-2024)
    rsi = wb['RSI']
    
    # Row mappings for RSI
    rsi_rows = {
        'service_cost': 4,
        'interest': 5,
        'benefit_changes': 6,
        'experience': 7,
        'assumption_changes': 8,
        'benefit_payments': 9,
        'net_change': 10,
        'tol_beginning': 12,
        'tol_ending': 14,
        'covered_payroll': 17,
        'liability_pct': 20,
        'discount_rates': 26,
        'mortality': 27,
        'trend': 28
    }
    
    cols = ['B', 'C', 'D', 'E', 'F', 'G', 'H']  # 2018-2024
    
    for attr, row in rsi_rows.items():
        values = []
        for col in cols:
            val = rsi[f'{col}{row}'].value
            values.append(val if val is not None else 0)
        setattr(data, f'rsi_{attr}', values)
    
    # OPEB Exp & Def - get current year values that become prior year
    opeb = wb['OPEB Exp & Def']
    
    # B7 in current year = B6 from prior year (current year assumption change becomes prior)
    # B24 in current year = B10 from prior year (current year experience becomes prior)
    data.opeb_assumption_change_prior = opeb['B6'].value or 0  # Prior file's B6 → Current file's B7
    data.opeb_experience_prior = opeb['B10'].value or 0  # Prior file's B10 → Current file's B24
    
    # Table7AmortDeferred - extract yellow cells
    # Column B = Total (round to int), Column C = ARSL (integer), Column AO = EOY Balance
    # The AO (EOY Balance) from prior year becomes AN (PY Balance) in current year (rounded)
    amort = wb['Table7AmortDeferred']
    
    # Experience section - in 2024 file, rows 13-19 contain years 2024-2018
    # Column C has the ARSL integers (5, 6, 7, 8, 9, 11, 12)
    exp_rows_2024 = {13: 2024, 14: 2023, 15: 2022, 16: 2021, 17: 2020, 18: 2019, 19: 2018}
    
    for src_row, year in exp_rows_2024.items():
        year_check = amort[f'A{src_row}'].value
        if year_check == year or str(year_check) == str(year):
            total = amort[f'B{src_row}'].value
            arsl = amort[f'C{src_row}'].value
            eoy_bal = amort[f'AO{src_row}'].value  # EOY becomes next year's PY Balance
            data.amort_experience_years[year] = {
                'total': round(total) if total else 0,
                'arsl': arsl,  # ARSL integer from column C
                'py_balance': round(eoy_bal) if eoy_bal else 0  # Use AO, rounded
            }
    
    # Assumption changes section - rows 23-29 in 2024 file contain years 2024-2018
    assump_rows_2024 = {23: 2024, 24: 2023, 25: 2022, 26: 2021, 27: 2020, 28: 2019, 29: 2018}
    
    for src_row, year in assump_rows_2024.items():
        year_check = amort[f'A{src_row}'].value
        if year_check == year or str(year_check) == str(year):
            total = amort[f'B{src_row}'].value
            arsl = amort[f'C{src_row}'].value
            eoy_bal = amort[f'AO{src_row}'].value  # EOY becomes next year's PY Balance
            data.amort_assumption_years[year] = {
                'total': round(total) if total else 0,
                'arsl': arsl,  # ARSL integer
                'py_balance': round(eoy_bal) if eoy_bal else 0  # Use AO, rounded
            }
    
    wb.close()
    return data


def populate_proval1(sheet, results: ValuationResults):
    """
    Populate the ProVal1 tab with valuation engine results.
    
    This is the primary data input sheet - all other sheets reference this.
    
    NOTE: B74 and C74 (BOY ARSL values) are NOT overwritten - they should
    retain their template values from the prior year valuation.
    """
    # Column B: BOY at old rate
    sheet['B6'] = results.active_count  # Use BOY counts (approximation)
    sheet['B7'] = results.active_count_fully_eligible
    sheet['B8'] = results.retiree_count
    sheet['B17'] = results.covered_payroll
    sheet['B18'] = results.total_salary_fully_eligible
    sheet['B19'] = results.tol_boy_old_rate
    sheet['B38'] = results.service_cost
    # B74 - DO NOT OVERWRITE - keep prior year's BOY ARSL
    sheet['B88'] = results.discount_rate_boy
    
    # Column C: BOY at new rate (for assumption change calculation)
    sheet['C6'] = results.active_count
    sheet['C7'] = results.active_count_fully_eligible
    sheet['C8'] = results.retiree_count
    sheet['C17'] = results.covered_payroll
    sheet['C18'] = results.total_salary_fully_eligible
    sheet['C19'] = results.tol_boy_new_rate
    sheet['C38'] = results.service_cost
    # C74 - DO NOT OVERWRITE - keep prior year's BOY ARSL at new rate
    sheet['C88'] = results.discount_rate_eoy
    
    # Column D: EOY baseline
    sheet['D6'] = results.active_count
    sheet['D7'] = results.active_count_fully_eligible
    sheet['D8'] = results.retiree_count
    sheet['D17'] = results.covered_payroll
    sheet['D18'] = results.total_salary_fully_eligible
    sheet['D19'] = results.tol_eoy_baseline
    sheet['D38'] = results.service_cost
    sheet['D74'] = results.avg_remaining_service_life
    sheet['D88'] = results.discount_rate_eoy
    
    # Column E: EOY discount +1%
    sheet['E6'] = results.active_count
    sheet['E8'] = results.retiree_count
    sheet['E17'] = results.covered_payroll
    sheet['E19'] = results.tol_eoy_disc_plus_1
    sheet['E74'] = results.avg_remaining_service_life
    sheet['E88'] = results.discount_rate_eoy + 0.01
    
    # Column F: EOY discount -1%
    sheet['F6'] = results.active_count
    sheet['F8'] = results.retiree_count
    sheet['F17'] = results.covered_payroll
    sheet['F19'] = results.tol_eoy_disc_minus_1
    sheet['F74'] = results.avg_remaining_service_life
    sheet['F88'] = results.discount_rate_eoy - 0.01
    
    # Column G: EOY trend baseline (same as D for liability)
    sheet['G6'] = results.active_count
    sheet['G8'] = results.retiree_count
    sheet['G17'] = results.covered_payroll
    sheet['G19'] = results.tol_eoy_baseline  # Same as baseline
    sheet['G74'] = results.avg_remaining_service_life
    
    # Column H: EOY trend +1%
    sheet['H6'] = results.active_count
    sheet['H8'] = results.retiree_count
    sheet['H17'] = results.covered_payroll
    sheet['H19'] = results.tol_eoy_trend_plus_1
    sheet['H74'] = results.avg_remaining_service_life
    
    # Column I: EOY trend -1%
    sheet['I6'] = results.active_count
    sheet['I8'] = results.retiree_count
    sheet['I17'] = results.covered_payroll
    sheet['I19'] = results.tol_eoy_trend_minus_1
    sheet['I74'] = results.avg_remaining_service_life


def populate_rsi_historical(sheet, prior_data: PriorYearData):
    """
    Populate RSI tab with historical data from prior year file.
    Columns B-H (2018-2024) are pasted as values.
    Column I (current year) uses formulas referencing Net OPEB.
    """
    cols = ['B', 'C', 'D', 'E', 'F', 'G', 'H']
    
    # Row 4: Service cost
    for i, col in enumerate(cols):
        if i < len(prior_data.rsi_service_cost):
            sheet[f'{col}4'] = prior_data.rsi_service_cost[i]
    
    # Row 5: Interest
    for i, col in enumerate(cols):
        if i < len(prior_data.rsi_interest):
            sheet[f'{col}5'] = prior_data.rsi_interest[i]
    
    # Row 6: Benefit changes
    for i, col in enumerate(cols):
        if i < len(prior_data.rsi_benefit_changes):
            sheet[f'{col}6'] = prior_data.rsi_benefit_changes[i]
    
    # Row 7: Experience
    for i, col in enumerate(cols):
        if i < len(prior_data.rsi_experience):
            sheet[f'{col}7'] = prior_data.rsi_experience[i]
    
    # Row 8: Assumption changes
    for i, col in enumerate(cols):
        if i < len(prior_data.rsi_assumption_changes):
            sheet[f'{col}8'] = prior_data.rsi_assumption_changes[i]
    
    # Row 9: Benefit payments
    for i, col in enumerate(cols):
        if i < len(prior_data.rsi_benefit_payments):
            sheet[f'{col}9'] = prior_data.rsi_benefit_payments[i]
    
    # Row 10: Net change
    for i, col in enumerate(cols):
        if i < len(prior_data.rsi_net_change):
            sheet[f'{col}10'] = prior_data.rsi_net_change[i]
    
    # Row 12: TOL beginning
    for i, col in enumerate(cols):
        if i < len(prior_data.rsi_tol_beginning):
            sheet[f'{col}12'] = prior_data.rsi_tol_beginning[i]
    
    # Row 14: TOL ending
    for i, col in enumerate(cols):
        if i < len(prior_data.rsi_tol_ending):
            sheet[f'{col}14'] = prior_data.rsi_tol_ending[i]
    
    # Row 17: Covered payroll
    for i, col in enumerate(cols):
        if i < len(prior_data.rsi_covered_payroll):
            sheet[f'{col}17'] = prior_data.rsi_covered_payroll[i]
    
    # Row 20: Liability %
    for i, col in enumerate(cols):
        if i < len(prior_data.rsi_liability_pct):
            sheet[f'{col}20'] = prior_data.rsi_liability_pct[i]
    
    # Row 26: Discount rates
    for i, col in enumerate(cols):
        if i < len(prior_data.rsi_discount_rates):
            sheet[f'{col}26'] = prior_data.rsi_discount_rates[i]
    
    # Row 27: Mortality
    for i, col in enumerate(cols):
        if i < len(prior_data.rsi_mortality):
            sheet[f'{col}27'] = prior_data.rsi_mortality[i]
    
    # Row 28: Trend
    for i, col in enumerate(cols):
        if i < len(prior_data.rsi_trend):
            sheet[f'{col}28'] = prior_data.rsi_trend[i]


def populate_opeb_exp_def(sheet, prior_data: PriorYearData):
    """
    Populate OPEB Exp & Def tab with prior year amortization values.
    Only B7 and B24 need to be populated from prior year.
    These are rounded to integers to match target formatting.
    """
    # Round to integers like the target file
    sheet['B7'] = round(prior_data.opeb_assumption_change_prior)
    sheet['B24'] = round(prior_data.opeb_experience_prior)


def populate_table7_amort(sheet, prior_data: PriorYearData):
    """
    Populate Table7AmortDeferred yellow cells from prior year data.
    
    Yellow cells to populate:
    - Column B: Total deferred amount
    - Column J: Amortization Years (ARSL - integer service life when established)
    - Column AN: Prior Year Balance
    
    For Experience section (rows 14-20, years 2024-2018)
    For Assumption section (rows 27-33, years 2024-2018)
    """
    # Experience section - rows 14-20 in 2025 file (years 2024-2018)
    exp_row_map = {2024: 14, 2023: 15, 2022: 16, 2021: 17, 2020: 18, 2019: 19, 2018: 20}
    
    for year, dest_row in exp_row_map.items():
        if year in prior_data.amort_experience_years:
            data = prior_data.amort_experience_years[year]
            if data.get('total') is not None:
                sheet[f'B{dest_row}'] = data['total']
            if data.get('arsl') is not None:
                sheet[f'J{dest_row}'] = data['arsl']  # Integer service life
            if data.get('py_balance') is not None:
                sheet[f'AN{dest_row}'] = data['py_balance']
    
    # Assumption section - rows 27-33 in 2025 file (years 2024-2018)
    assump_row_map = {2024: 27, 2023: 28, 2022: 29, 2021: 30, 2020: 31, 2019: 32, 2018: 33}
    
    for year, dest_row in assump_row_map.items():
        if year in prior_data.amort_assumption_years:
            data = prior_data.amort_assumption_years[year]
            if data.get('total') is not None:
                sheet[f'B{dest_row}'] = data['total']
            if data.get('arsl') is not None:
                sheet[f'J{dest_row}'] = data['arsl']  # Integer service life
            if data.get('py_balance') is not None:
                sheet[f'AN{dest_row}'] = data['py_balance']


def remove_highlighting(wb):
    """Remove all yellow highlighting and conditional formatting from workbook."""
    no_fill = PatternFill(fill_type=None)
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet.conditional_formatting = ConditionalFormattingList()
        
        for row in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                if cell.fill and cell.fill.fgColor:
                    color = str(cell.fill.fgColor.rgb) if cell.fill.fgColor.rgb else ''
                    if 'FFFF' in color:  # Yellow
                        cell.fill = no_fill


def remove_comments(wb):
    """Remove all comments/notes from workbook."""
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                cell.comment = None


def populate_gasb75_disclosure(
    template_path: str,
    output_path: str,
    results: ValuationResults,
    prior_year_path: Optional[str] = None,
    remove_yellow: bool = True,
    remove_notes: bool = True
) -> str:
    """
    Main function to populate GASB 75 disclosure workbook.
    
    Args:
        template_path: Path to blank GASB 75 template Excel file
        output_path: Path for output populated Excel file
        results: ValuationResults from the valuation engine
        prior_year_path: Path to prior year's GASB 75 file (for RSI, amortization)
        remove_yellow: Remove yellow highlighting from output
        remove_notes: Remove cell comments from output
        
    Returns:
        Path to the output file
    """
    # Load template
    wb = load_workbook(template_path)
    
    # Extract prior year data if provided
    prior_data = None
    if prior_year_path:
        prior_data = extract_prior_year_data(prior_year_path)
    
    # 1. Populate ProVal1 - the primary data input tab
    populate_proval1(wb['ProVal1'], results)
    
    # 2. Populate RSI historical columns if prior year data available
    if prior_data:
        populate_rsi_historical(wb['RSI'], prior_data)
        
        # 3. Populate OPEB Exp & Def prior year values
        populate_opeb_exp_def(wb['OPEB Exp & Def'], prior_data)
        
        # 4. Populate Table7AmortDeferred yellow cells
        populate_table7_amort(wb['Table7AmortDeferred'], prior_data)
    
    # Remove formatting as requested
    if remove_yellow:
        remove_highlighting(wb)
    
    if remove_notes:
        remove_comments(wb)
    
    # Save output
    wb.save(output_path)
    wb.close()
    
    return output_path


# Convenience function for integration with valuation engine
def create_gasb75_disclosure_from_engine(
    engine_results: Dict[str, Any],
    template_path: str,
    output_path: str,
    prior_year_path: Optional[str] = None,
    client_name: str = "Client",
    measurement_date: date = None,
    prior_measurement_date: date = None
) -> str:
    """
    Create GASB 75 disclosure from valuation engine output dictionary.
    
    This is the interface function for the opeb-full-valuation pipeline.
    
    Args:
        engine_results: Dictionary with keys matching ValuationResults fields
        template_path: Path to GASB 75 template
        output_path: Output path
        prior_year_path: Prior year GASB 75 file
        client_name: Client name for the disclosure
        measurement_date: EOY measurement date
        prior_measurement_date: BOY measurement date
        
    Returns:
        Path to populated disclosure file
    """
    # Build ValuationResults from engine output
    results = ValuationResults(
        client_name=client_name,
        measurement_date=measurement_date or date.today(),
        prior_measurement_date=prior_measurement_date or date.today(),
        active_count=engine_results.get('active_count', 0),
        retiree_count=engine_results.get('retiree_count', 0),
        covered_payroll=engine_results.get('covered_payroll', 0),
        tol_boy_old_rate=engine_results.get('tol_boy_old_rate', 0),
        tol_boy_new_rate=engine_results.get('tol_boy_new_rate', 0),
        tol_eoy_baseline=engine_results.get('tol_eoy_baseline', 0),
        tol_eoy_disc_plus_1=engine_results.get('tol_eoy_disc_plus_1', 0),
        tol_eoy_disc_minus_1=engine_results.get('tol_eoy_disc_minus_1', 0),
        tol_eoy_trend_plus_1=engine_results.get('tol_eoy_trend_plus_1', 0),
        tol_eoy_trend_minus_1=engine_results.get('tol_eoy_trend_minus_1', 0),
        service_cost=engine_results.get('service_cost', 0),
        discount_rate_boy=engine_results.get('discount_rate_boy', 0),
        discount_rate_eoy=engine_results.get('discount_rate_eoy', 0),
        avg_remaining_service_life=engine_results.get('avg_remaining_service_life', 5.0),
        benefit_payments=engine_results.get('benefit_payments', 0)
    )
    
    return populate_gasb75_disclosure(
        template_path=template_path,
        output_path=output_path,
        results=results,
        prior_year_path=prior_year_path,
        remove_yellow=True,
        remove_notes=True
    )


if __name__ == '__main__':
    # Test with West Florida Planning data
    print("GASB 75 Disclosure Population Module")
    print("=" * 50)
    print("Use populate_gasb75_disclosure() or")
    print("create_gasb75_disclosure_from_engine() to populate disclosures.")
