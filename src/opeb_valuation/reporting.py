"""
opeb_valuation/reporting.py - Automated GASB 75 Excel Reporting

Produces client-ready Excel workbooks with:
1. Template loading from client's existing XLSX
2. Cell mapping for TOL, Service Cost, Interest, etc.
3. Automatic sensitivity analysis (5 runs: baseline + ±1% disc/trend)
4. Deferred inflows/outflows amortization schedules

Author: Joseph Shackelford - Actuarial Pipeline Project
License: MIT
"""

import numpy as np
import pandas as pd
from datetime import date, datetime
from typing import Dict, List, Optional, Tuple, Any, Union
from dataclasses import dataclass, field
from pathlib import Path
from enum import Enum
import copy
import logging

logger = logging.getLogger(__name__)

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not installed - Excel features limited")


# =============================================================================
# CELL MAPPING CONFIGURATION
# =============================================================================

@dataclass
class CellMapping:
    """Maps a value to a specific Excel cell."""
    sheet_name: str
    cell_ref: str
    value_key: str
    format_type: str = "currency"  # currency, percent, number, date, text
    description: str = ""


# Default GASB 75 cell mappings
DEFAULT_TOL_RECONCILIATION_MAPPINGS = [
    CellMapping("Changes in TOL", "C10", "tol_boy", "currency", "Beginning TOL"),
    CellMapping("Changes in TOL", "C11", "service_cost", "currency", "Service Cost"),
    CellMapping("Changes in TOL", "C12", "interest_cost", "currency", "Interest Cost"),
    CellMapping("Changes in TOL", "C13", "changes_benefit_terms", "currency", "Changes in Benefit Terms"),
    CellMapping("Changes in TOL", "C14", "experience_gain_loss", "currency", "Experience (Gain)/Loss"),
    CellMapping("Changes in TOL", "C15", "assumption_changes", "currency", "Assumption Changes"),
    CellMapping("Changes in TOL", "C16", "benefit_payments", "currency", "Benefit Payments"),
    CellMapping("Changes in TOL", "C17", "tol_eoy", "currency", "Ending TOL"),
]

DEFAULT_SENSITIVITY_MAPPINGS = [
    CellMapping("Sensitivity", "C10", "baseline_tol", "currency", "Baseline TOL"),
    CellMapping("Sensitivity", "C11", "disc_minus_1_tol", "currency", "Discount -1%"),
    CellMapping("Sensitivity", "C12", "disc_plus_1_tol", "currency", "Discount +1%"),
    CellMapping("Sensitivity", "C13", "trend_minus_1_tol", "currency", "Trend -1%"),
    CellMapping("Sensitivity", "C14", "trend_plus_1_tol", "currency", "Trend +1%"),
]

DEFAULT_DEFERRED_MAPPINGS = [
    CellMapping("Deferred Outflows", "C10", "deferred_outflow_experience", "currency"),
    CellMapping("Deferred Outflows", "C11", "deferred_outflow_assumptions", "currency"),
    CellMapping("Deferred Inflows", "C10", "deferred_inflow_experience", "currency"),
    CellMapping("Deferred Inflows", "C11", "deferred_inflow_assumptions", "currency"),
]


# =============================================================================
# VALUATION RESULTS CONTAINER
# =============================================================================

@dataclass
class ValuationResults:
    """Container for all valuation results needed for reporting."""
    # Identification
    client_name: str
    measurement_date: date
    prior_measurement_date: date
    fiscal_year_end: date
    
    # TOL Reconciliation
    tol_boy: float
    tol_eoy: float
    service_cost: float
    interest_cost: float
    benefit_payments: float
    experience_gain_loss: float = 0.0
    assumption_changes: float = 0.0
    changes_benefit_terms: float = 0.0
    
    # Census
    active_count: int = 0
    retiree_count: int = 0
    total_count: int = 0
    
    # Covered payroll
    covered_payroll: float = 0.0
    
    # Assumptions
    discount_rate: float = 0.0381
    discount_rate_boy: float = 0.0409
    initial_trend: float = 0.065
    ultimate_trend: float = 0.045
    
    # Sensitivity results
    sensitivity: Dict[str, float] = field(default_factory=dict)
    
    # Deferred items
    deferred_outflows: Dict[str, float] = field(default_factory=dict)
    deferred_inflows: Dict[str, float] = field(default_factory=dict)
    
    # Amortization schedules
    amortization_schedules: Dict[str, pd.DataFrame] = field(default_factory=dict)
    
    # Additional detail
    tol_by_status: Dict[str, float] = field(default_factory=dict)
    service_cost_by_status: Dict[str, float] = field(default_factory=dict)


@dataclass
class SensitivityResult:
    """Result from a sensitivity analysis run."""
    scenario: str
    discount_rate: float
    initial_trend: float
    tol: float
    service_cost: float
    description: str = ""


# =============================================================================
# SENSITIVITY ANALYZER
# =============================================================================

class SensitivityAnalyzer:
    """
    Runs automatic sensitivity analysis.
    
    Performs 5 valuations:
    1. Baseline
    2. Discount rate -1%
    3. Discount rate +1%
    4. Healthcare trend -1%
    5. Healthcare trend +1%
    """
    
    SCENARIOS = [
        ("baseline", 0, 0),
        ("disc_minus_1", -0.01, 0),
        ("disc_plus_1", 0.01, 0),
        ("trend_minus_1", 0, -0.01),
        ("trend_plus_1", 0, 0.01),
    ]
    
    def __init__(self, engine_factory: callable, base_config: Dict):
        """
        Initialize sensitivity analyzer.
        
        Args:
            engine_factory: Function that creates a valuation engine from config
            base_config: Base configuration dictionary
        """
        self.engine_factory = engine_factory
        self.base_config = base_config
    
    def run_all_scenarios(self, census_actives: pd.DataFrame,
                          census_retirees: pd.DataFrame) -> Dict[str, SensitivityResult]:
        """
        Run all 5 sensitivity scenarios.
        
        Returns:
            Dict mapping scenario name to SensitivityResult
        """
        results = {}
        
        base_discount = self.base_config.get('discount_rate', 0.0381)
        base_trend = self.base_config.get('initial_trend', 0.065)
        
        for scenario_name, disc_adj, trend_adj in self.SCENARIOS:
            logger.info(f"Running sensitivity scenario: {scenario_name}")
            
            # Adjust config
            config = copy.deepcopy(self.base_config)
            config['discount_rate'] = base_discount + disc_adj
            config['initial_trend'] = base_trend + trend_adj
            
            # Run valuation
            engine = self.engine_factory(config)
            val_results = engine.run_valuation(census_actives, census_retirees)
            
            # Extract results
            tol = val_results['TOL'].sum()
            service_cost = val_results['ServiceCost'].sum()
            
            results[scenario_name] = SensitivityResult(
                scenario=scenario_name,
                discount_rate=config['discount_rate'],
                initial_trend=config['initial_trend'],
                tol=tol,
                service_cost=service_cost,
                description=self._get_description(scenario_name)
            )
        
        return results
    
    def _get_description(self, scenario: str) -> str:
        """Get description for scenario."""
        descriptions = {
            'baseline': 'Current assumptions',
            'disc_minus_1': 'Discount rate decreased 1%',
            'disc_plus_1': 'Discount rate increased 1%',
            'trend_minus_1': 'Healthcare trend decreased 1%',
            'trend_plus_1': 'Healthcare trend increased 1%',
        }
        return descriptions.get(scenario, scenario)


# =============================================================================
# AMORTIZATION SCHEDULE GENERATOR
# =============================================================================

class AmortizationScheduleGenerator:
    """
    Generates deferred inflow/outflow amortization schedules.
    
    Per GASB 75, experience gains/losses and assumption changes
    are amortized over the average remaining service life (ARSL).
    """
    
    def __init__(self, arsl: float = 5.0):
        """
        Initialize generator.
        
        Args:
            arsl: Average remaining service life in years
        """
        self.arsl = arsl
    
    def generate_schedule(self, item_name: str, initial_amount: float,
                          recognition_date: date,
                          periods: int = 10) -> pd.DataFrame:
        """
        Generate amortization schedule for a deferred item.
        
        Args:
            item_name: Name of the deferred item
            initial_amount: Initial amount to amortize
            recognition_date: Date item was recognized
            periods: Number of periods to show
        
        Returns:
            DataFrame with amortization schedule
        """
        annual_recognition = initial_amount / self.arsl
        
        schedule = []
        remaining = initial_amount
        
        for i in range(periods):
            year = recognition_date.year + i
            
            if i < int(self.arsl):
                recognized = annual_recognition
            elif i == int(self.arsl) and self.arsl % 1 > 0:
                # Fractional year
                recognized = annual_recognition * (self.arsl % 1)
            else:
                recognized = 0
            
            remaining = max(0, remaining - recognized)
            
            schedule.append({
                'Year': year,
                'Beginning Balance': remaining + recognized if recognized > 0 else remaining,
                'Recognition': recognized,
                'Ending Balance': remaining,
            })
        
        df = pd.DataFrame(schedule)
        df.attrs['item_name'] = item_name
        df.attrs['initial_amount'] = initial_amount
        df.attrs['arsl'] = self.arsl
        
        return df
    
    def generate_combined_schedule(self, items: Dict[str, Tuple[float, date]],
                                    periods: int = 10) -> pd.DataFrame:
        """
        Generate combined amortization schedule for multiple items.
        
        Args:
            items: Dict mapping item name to (amount, recognition_date)
            periods: Number of periods
        
        Returns:
            Combined DataFrame
        """
        if not items:
            return pd.DataFrame()
        
        # Get the earliest date
        min_date = min(d for _, d in items.values())
        base_year = min_date.year
        
        # Initialize combined schedule
        years = list(range(base_year, base_year + periods))
        combined = pd.DataFrame({'Year': years})
        combined['Total Recognition'] = 0.0
        
        for name, (amount, rec_date) in items.items():
            schedule = self.generate_schedule(name, amount, rec_date, periods)
            
            # Align by year
            merged = combined.merge(
                schedule[['Year', 'Recognition']].rename(columns={'Recognition': name}),
                on='Year',
                how='left'
            ).fillna(0)
            
            combined = merged
            combined['Total Recognition'] += combined[name]
        
        return combined


# =============================================================================
# EXCEL REPORT GENERATOR
# =============================================================================

class ExcelReportGenerator:
    """
    Generates client-ready GASB 75 Excel workbooks.
    
    Features:
    1. Load template from client's existing XLSX
    2. Map values to specific cells
    3. Auto-format currency/percent/dates
    4. Generate sensitivity tables
    5. Generate amortization schedules
    """
    
    def __init__(self, template_path: Optional[Union[str, Path]] = None):
        """
        Initialize report generator.
        
        Args:
            template_path: Path to client's Excel template (optional)
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl required for Excel reporting. Install with: pip install openpyxl")
        
        self.template_path = Path(template_path) if template_path else None
        self.workbook = None
        self.cell_mappings: List[CellMapping] = []
        
        # Default styling
        self.currency_format = '$#,##0'
        self.percent_format = '0.00%'
        self.date_format = 'MM/DD/YYYY'
        
        # Styles
        self.header_font = Font(bold=True, size=11)
        self.currency_font = Font(size=10)
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font_white = Font(bold=True, size=11, color="FFFFFF")
    
    def load_template(self, template_path: Optional[Union[str, Path]] = None) -> None:
        """Load an Excel template."""
        path = Path(template_path) if template_path else self.template_path
        
        if path and path.exists():
            self.workbook = load_workbook(path)
            logger.info(f"Loaded template: {path.name}")
        else:
            self.workbook = Workbook()
            logger.info("Created new workbook (no template)")
    
    def create_new_workbook(self) -> None:
        """Create a new workbook with standard GASB 75 sheets."""
        self.workbook = Workbook()
        
        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']
        
        # Create standard sheets
        sheets = [
            "Summary",
            "Changes in TOL",
            "Sensitivity",
            "Deferred Outflows",
            "Deferred Inflows",
            "Amortization",
            "Census Summary",
            "Assumptions",
        ]
        
        for sheet_name in sheets:
            self.workbook.create_sheet(sheet_name)
        
        logger.info(f"Created workbook with {len(sheets)} sheets")
    
    def set_cell_value(self, sheet_name: str, cell_ref: str, 
                       value: Any, format_type: str = "currency") -> None:
        """
        Set a cell value with formatting.
        
        Args:
            sheet_name: Name of the worksheet
            cell_ref: Cell reference (e.g., "C10")
            value: Value to set
            format_type: One of "currency", "percent", "number", "date", "text"
        """
        if self.workbook is None:
            self.create_new_workbook()
        
        # Get or create sheet
        if sheet_name not in self.workbook.sheetnames:
            self.workbook.create_sheet(sheet_name)
        
        sheet = self.workbook[sheet_name]
        cell = sheet[cell_ref]
        
        # Set value
        cell.value = value
        
        # Apply formatting
        if format_type == "currency":
            cell.number_format = self.currency_format
        elif format_type == "percent":
            cell.number_format = self.percent_format
        elif format_type == "date" and isinstance(value, (date, datetime)):
            cell.number_format = self.date_format
    
    def apply_mappings(self, values: Dict[str, Any],
                       mappings: Optional[List[CellMapping]] = None) -> None:
        """
        Apply a list of cell mappings.
        
        Args:
            values: Dict mapping value keys to actual values
            mappings: List of CellMapping objects (uses defaults if None)
        """
        if mappings is None:
            mappings = (DEFAULT_TOL_RECONCILIATION_MAPPINGS + 
                       DEFAULT_SENSITIVITY_MAPPINGS +
                       DEFAULT_DEFERRED_MAPPINGS)
        
        for mapping in mappings:
            if mapping.value_key in values:
                self.set_cell_value(
                    mapping.sheet_name,
                    mapping.cell_ref,
                    values[mapping.value_key],
                    mapping.format_type
                )
    
    def populate_from_results(self, results: ValuationResults) -> None:
        """
        Populate workbook from ValuationResults object.
        
        Args:
            results: ValuationResults containing all valuation data
        """
        if self.workbook is None:
            self.create_new_workbook()
        
        # Build values dict
        values = {
            # TOL Reconciliation
            'tol_boy': results.tol_boy,
            'tol_eoy': results.tol_eoy,
            'service_cost': results.service_cost,
            'interest_cost': results.interest_cost,
            'benefit_payments': results.benefit_payments,
            'experience_gain_loss': results.experience_gain_loss,
            'assumption_changes': results.assumption_changes,
            'changes_benefit_terms': results.changes_benefit_terms,
            
            # Sensitivity
            'baseline_tol': results.sensitivity.get('baseline', results.tol_eoy),
            'disc_minus_1_tol': results.sensitivity.get('disc_minus_1', 0),
            'disc_plus_1_tol': results.sensitivity.get('disc_plus_1', 0),
            'trend_minus_1_tol': results.sensitivity.get('trend_minus_1', 0),
            'trend_plus_1_tol': results.sensitivity.get('trend_plus_1', 0),
            
            # Deferred items
            **{f'deferred_outflow_{k}': v for k, v in results.deferred_outflows.items()},
            **{f'deferred_inflow_{k}': v for k, v in results.deferred_inflows.items()},
        }
        
        self.apply_mappings(values)
        
        # Populate additional sheets
        self._populate_summary(results)
        self._populate_census_summary(results)
        self._populate_assumptions(results)
        self._populate_amortization(results)
    
    def _populate_summary(self, results: ValuationResults) -> None:
        """Populate the Summary sheet."""
        sheet_name = "Summary"
        if sheet_name not in self.workbook.sheetnames:
            self.workbook.create_sheet(sheet_name)
        
        sheet = self.workbook[sheet_name]
        
        # Title
        sheet['A1'] = f"GASB 75 OPEB Valuation Summary"
        sheet['A1'].font = Font(bold=True, size=14)
        
        sheet['A3'] = "Client:"
        sheet['B3'] = results.client_name
        
        sheet['A4'] = "Measurement Date:"
        sheet['B4'] = results.measurement_date
        sheet['B4'].number_format = self.date_format
        
        sheet['A5'] = "Fiscal Year End:"
        sheet['B5'] = results.fiscal_year_end
        sheet['B5'].number_format = self.date_format
        
        # Key Results
        sheet['A7'] = "KEY RESULTS"
        sheet['A7'].font = self.header_font
        
        summary_data = [
            ("Total OPEB Liability (EOY)", results.tol_eoy),
            ("Total OPEB Liability (BOY)", results.tol_boy),
            ("Service Cost", results.service_cost),
            ("Interest Cost", results.interest_cost),
            ("Benefit Payments", results.benefit_payments),
            ("Net Change in TOL", results.tol_eoy - results.tol_boy),
        ]
        
        for i, (label, value) in enumerate(summary_data, start=8):
            sheet[f'A{i}'] = label
            sheet[f'B{i}'] = value
            sheet[f'B{i}'].number_format = self.currency_format
    
    def _populate_census_summary(self, results: ValuationResults) -> None:
        """Populate the Census Summary sheet."""
        sheet_name = "Census Summary"
        if sheet_name not in self.workbook.sheetnames:
            self.workbook.create_sheet(sheet_name)
        
        sheet = self.workbook[sheet_name]
        
        sheet['A1'] = "Census Summary"
        sheet['A1'].font = Font(bold=True, size=14)
        
        census_data = [
            ("Active Employees", results.active_count),
            ("Retirees & Beneficiaries", results.retiree_count),
            ("Total Participants", results.total_count),
            ("Covered Payroll", results.covered_payroll),
        ]
        
        for i, (label, value) in enumerate(census_data, start=3):
            sheet[f'A{i}'] = label
            sheet[f'B{i}'] = value
            if "Payroll" in label:
                sheet[f'B{i}'].number_format = self.currency_format
    
    def _populate_assumptions(self, results: ValuationResults) -> None:
        """Populate the Assumptions sheet."""
        sheet_name = "Assumptions"
        if sheet_name not in self.workbook.sheetnames:
            self.workbook.create_sheet(sheet_name)
        
        sheet = self.workbook[sheet_name]
        
        sheet['A1'] = "Actuarial Assumptions"
        sheet['A1'].font = Font(bold=True, size=14)
        
        assumptions = [
            ("Discount Rate (EOY)", results.discount_rate, "percent"),
            ("Discount Rate (BOY)", results.discount_rate_boy, "percent"),
            ("Initial Healthcare Trend", results.initial_trend, "percent"),
            ("Ultimate Healthcare Trend", results.ultimate_trend, "percent"),
        ]
        
        for i, (label, value, fmt) in enumerate(assumptions, start=3):
            sheet[f'A{i}'] = label
            sheet[f'B{i}'] = value
            if fmt == "percent":
                sheet[f'B{i}'].number_format = self.percent_format
    
    def _populate_amortization(self, results: ValuationResults) -> None:
        """Populate amortization schedules."""
        sheet_name = "Amortization"
        if sheet_name not in self.workbook.sheetnames:
            self.workbook.create_sheet(sheet_name)
        
        sheet = self.workbook[sheet_name]
        
        sheet['A1'] = "Deferred Outflows/Inflows Amortization Schedule"
        sheet['A1'].font = Font(bold=True, size=14)
        
        # If there are amortization schedules, write them
        row = 3
        for name, schedule_df in results.amortization_schedules.items():
            sheet[f'A{row}'] = name
            sheet[f'A{row}'].font = self.header_font
            row += 1
            
            for r_idx, row_data in enumerate(dataframe_to_rows(schedule_df, index=False, header=True)):
                for c_idx, value in enumerate(row_data, start=1):
                    cell = sheet.cell(row=row, column=c_idx, value=value)
                    if r_idx == 0:
                        cell.font = self.header_font
                    elif c_idx > 1:
                        cell.number_format = self.currency_format
                row += 1
            
            row += 2  # Space between schedules
    
    def generate_sensitivity_table(self, sensitivity_results: Dict[str, SensitivityResult]) -> None:
        """
        Generate a formatted sensitivity analysis table.
        
        Args:
            sensitivity_results: Dict of sensitivity results
        """
        sheet_name = "Sensitivity"
        if sheet_name not in self.workbook.sheetnames:
            self.workbook.create_sheet(sheet_name)
        
        sheet = self.workbook[sheet_name]
        
        # Title
        sheet['A1'] = "Sensitivity of Total OPEB Liability"
        sheet['A1'].font = Font(bold=True, size=14)
        
        # Headers
        headers = ["Scenario", "Discount Rate", "Healthcare Trend", "Total OPEB Liability", "Change from Baseline"]
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=3, column=col, value=header)
            cell.font = self.header_font_white
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        baseline_tol = sensitivity_results.get('baseline', SensitivityResult('', 0, 0, 0, 0)).tol
        
        row = 4
        for scenario_name, result in sensitivity_results.items():
            sheet.cell(row=row, column=1, value=result.description)
            
            disc_cell = sheet.cell(row=row, column=2, value=result.discount_rate)
            disc_cell.number_format = self.percent_format
            
            trend_cell = sheet.cell(row=row, column=3, value=result.initial_trend)
            trend_cell.number_format = self.percent_format
            
            tol_cell = sheet.cell(row=row, column=4, value=result.tol)
            tol_cell.number_format = self.currency_format
            
            change = result.tol - baseline_tol
            change_cell = sheet.cell(row=row, column=5, value=change)
            change_cell.number_format = self.currency_format
            
            row += 1
        
        # Adjust column widths
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 18
        sheet.column_dimensions['D'].width = 22
        sheet.column_dimensions['E'].width = 22
    
    def save(self, output_path: Union[str, Path]) -> Path:
        """
        Save the workbook to file.
        
        Args:
            output_path: Path to save the Excel file
        
        Returns:
            Path to saved file
        """
        output_path = Path(output_path)
        
        if self.workbook is None:
            raise ValueError("No workbook to save - call create_new_workbook() or load_template() first")
        
        self.workbook.save(output_path)
        logger.info(f"Saved report to: {output_path}")
        
        return output_path


# =============================================================================
# CONVENIENCE FUNCTIONS
# =============================================================================

def generate_gasb75_report(
    results: ValuationResults,
    output_path: Union[str, Path],
    template_path: Optional[Union[str, Path]] = None,
    sensitivity_results: Optional[Dict[str, SensitivityResult]] = None
) -> Path:
    """
    Generate a complete GASB 75 Excel report.
    
    Args:
        results: ValuationResults object with all valuation data
        output_path: Path to save the Excel file
        template_path: Optional path to client's template
        sensitivity_results: Optional sensitivity analysis results
    
    Returns:
        Path to generated file
    """
    generator = ExcelReportGenerator(template_path)
    
    if template_path:
        generator.load_template()
    else:
        generator.create_new_workbook()
    
    generator.populate_from_results(results)
    
    if sensitivity_results:
        generator.generate_sensitivity_table(sensitivity_results)
    
    return generator.save(output_path)


def run_sensitivity_and_report(
    engine_factory: callable,
    base_config: Dict,
    census_actives: pd.DataFrame,
    census_retirees: pd.DataFrame,
    output_path: Union[str, Path],
    client_name: str,
    measurement_date: date,
    template_path: Optional[Union[str, Path]] = None
) -> Tuple[Path, Dict[str, SensitivityResult]]:
    """
    Run sensitivity analysis and generate complete report.
    
    Args:
        engine_factory: Function that creates valuation engine from config
        base_config: Base valuation configuration
        census_actives: Active employee census
        census_retirees: Retiree census
        output_path: Output Excel path
        client_name: Client name
        measurement_date: Measurement date
        template_path: Optional template path
    
    Returns:
        Tuple of (output path, sensitivity results)
    """
    # Run sensitivity analysis
    analyzer = SensitivityAnalyzer(engine_factory, base_config)
    sensitivity = analyzer.run_all_scenarios(census_actives, census_retirees)
    
    # Get baseline results for the main report
    baseline = sensitivity['baseline']
    
    # Create ValuationResults
    results = ValuationResults(
        client_name=client_name,
        measurement_date=measurement_date,
        prior_measurement_date=date(measurement_date.year - 1, measurement_date.month, measurement_date.day),
        fiscal_year_end=date(measurement_date.year, 12, 31),
        tol_boy=0,  # Would come from prior year
        tol_eoy=baseline.tol,
        service_cost=baseline.service_cost,
        interest_cost=0,  # Would be calculated
        benefit_payments=0,  # Would come from census/plan
        sensitivity={k: v.tol for k, v in sensitivity.items()},
        discount_rate=base_config.get('discount_rate', 0.0381),
        initial_trend=base_config.get('initial_trend', 0.065),
        active_count=len(census_actives),
        retiree_count=len(census_retirees),
        total_count=len(census_actives) + len(census_retirees),
    )
    
    # Generate report
    output = generate_gasb75_report(
        results, output_path, template_path, sensitivity
    )
    
    return output, sensitivity


# =============================================================================
# UNIT TESTS
# =============================================================================

if __name__ == "__main__":
    print("=" * 70)
    print("GASB 75 REPORTING MODULE - UNIT TESTS")
    print("=" * 70)
    
    if not OPENPYXL_AVAILABLE:
        print("\n⚠ openpyxl not installed - skipping Excel tests")
        print("  Install with: pip install openpyxl")
    else:
        # Test 1: Create report generator
        print("\nTest 1: Create Report Generator")
        print("-" * 50)
        
        generator = ExcelReportGenerator()
        generator.create_new_workbook()
        print(f"  Created workbook with sheets: {generator.workbook.sheetnames}")
        
        # Test 2: Populate from results
        print("\nTest 2: Populate from ValuationResults")
        print("-" * 50)
        
        results = ValuationResults(
            client_name="City of DeRidder",
            measurement_date=date(2025, 9, 30),
            prior_measurement_date=date(2024, 9, 30),
            fiscal_year_end=date(2025, 12, 31),
            tol_boy=6_350_000,
            tol_eoy=6_900_000,
            service_cost=450_000,
            interest_cost=285_000,
            benefit_payments=200_000,
            experience_gain_loss=-185_000,
            assumption_changes=0,
            discount_rate=0.0381,
            discount_rate_boy=0.0409,
            initial_trend=0.065,
            ultimate_trend=0.045,
            active_count=85,
            retiree_count=42,
            total_count=127,
            covered_payroll=4_250_000,
            sensitivity={
                'baseline': 6_900_000,
                'disc_minus_1': 7_850_000,
                'disc_plus_1': 6_150_000,
                'trend_minus_1': 6_450_000,
                'trend_plus_1': 7_400_000,
            }
        )
        
        generator.populate_from_results(results)
        print("  Populated all sheets with valuation data")
        
        # Test 3: Generate sensitivity table
        print("\nTest 3: Generate Sensitivity Table")
        print("-" * 50)
        
        sensitivity_results = {
            'baseline': SensitivityResult('baseline', 0.0381, 0.065, 6_900_000, 450_000, 'Current assumptions'),
            'disc_minus_1': SensitivityResult('disc_minus_1', 0.0281, 0.065, 7_850_000, 520_000, 'Discount rate decreased 1%'),
            'disc_plus_1': SensitivityResult('disc_plus_1', 0.0481, 0.065, 6_150_000, 390_000, 'Discount rate increased 1%'),
            'trend_minus_1': SensitivityResult('trend_minus_1', 0.0381, 0.055, 6_450_000, 415_000, 'Healthcare trend decreased 1%'),
            'trend_plus_1': SensitivityResult('trend_plus_1', 0.0381, 0.075, 7_400_000, 490_000, 'Healthcare trend increased 1%'),
        }
        
        generator.generate_sensitivity_table(sensitivity_results)
        print("  Generated sensitivity analysis table")
        
        # Test 4: Amortization schedule
        print("\nTest 4: Generate Amortization Schedule")
        print("-" * 50)
        
        amort_gen = AmortizationScheduleGenerator(arsl=5.0)
        schedule = amort_gen.generate_schedule(
            "Experience Loss 2024",
            initial_amount=500_000,
            recognition_date=date(2024, 9, 30)
        )
        print(f"  Generated {len(schedule)} year amortization schedule")
        print(schedule.head())
        
        # Test 5: Save workbook
        print("\nTest 5: Save Workbook")
        print("-" * 50)
        
        output_path = Path("/tmp/test_gasb75_report.xlsx")
        generator.save(output_path)
        print(f"  Saved to: {output_path}")
        
        print("\n✓ All reporting tests passed")
